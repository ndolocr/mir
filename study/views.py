# import xlwt using: pip install xlwt
#importing XlsxWriter: pip install XlsxWriter
# Importing openpyxl: pip install openpyxl
# Import Django rest Framework: pip install djangorestframework
import io
import os
import csv
import xlwt
import openpyxl

from io import StringIO
from zipfile import ZipFile

from study.models import Tag
from study.models import Study
from study.models import Theme
from study.models import Region
from study.models import Country
from study.models import Quality
from study.models import Category
from study.models import Resource
from study.models import SubTheme
from study.models import SubCategory

from django.conf import settings
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect

from xlsxwriter.workbook import Workbook
# Create your views here.

'''	Back End views'''


'''
	Starts Upload Category
'''
def category_upload(request):
	upload_template = "study/category_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/category_confirm.html", {"excel_data":excel_data})


def category_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Category.objects.update_or_create(
					category_name = cell,
				)
		return redirect('admin:index')


def download_category(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Category.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Category Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.category_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=category_list.xlsx"

	output.close()

	return response
	

'''
	End Upload Category
'''

'''
	Start Upload Country Information
'''

def country_upload(request):
	upload_template = "study/country_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/country_confirm.html", {"excel_data":excel_data})


def country_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Country.objects.update_or_create(
					country_name = cell,
				)
		return redirect('admin:index')

def download_country(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Country.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Country Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.country_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=country_list.xlsx"

	output.close()

	return response
	
'''
	End Upload Country Information
'''


'''
	Start Upload Quality Information
'''

def quality_upload(request):
	upload_template = "study/quality_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/quality_confirm.html", {"excel_data":excel_data})


def quality_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Quality.objects.update_or_create(
					quality_name = cell,
				)
		return redirect('admin:index')


def download_quality(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Quality.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Quality Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.quality_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=quality_list.xlsx"

	output.close()

	return response
	

'''
	End Upload Quality Information
'''


'''
	Start Upload Region Information
'''

def region_upload(request):
	upload_template = "study/region_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/region_confirm.html", {"excel_data":excel_data})


def region_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Region.objects.update_or_create(
					region_name = cell,
				)
		return redirect('admin:index')


def download_region(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Region.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Region Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.region_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=region_list.xlsx"

	output.close()

	return response
	
'''
	End Upload Region Information
'''


'''
	Start Upload Resource Information
'''

def resource_upload(request):
	upload_template = "study/resource_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/resource_confirm.html", {"excel_data":excel_data})


def resource_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Resource.objects.update_or_create(
					resource_name = cell,
				)
		return redirect('admin:index')


def download_resource(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Resource.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Resource Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.resource_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=resource_list.xlsx"

	output.close()

	return response

'''
	End Upload Resource Information
'''

'''
	Start Upload Theme Information
'''

def theme_upload(request):
	upload_template = "study/theme_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/theme_confirm.html", {"excel_data":excel_data})


def theme_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Theme.objects.update_or_create(
					theme_name = cell,
				)
		return redirect('admin:index')


def download_theme(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Theme.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 45)
	
	#column header names
	columns = ['Theme Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.theme_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=theme_list.xlsx"

	output.close()

	return response


'''
	End Upload Theme Information
'''

'''
	Start Upload Tags Information
'''

def tag_upload(request):
	upload_template = "study/tag_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/tag_confirm.html", {"excel_data":excel_data})


def tag_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			for cell in row:
				_, object_instance = Tag.objects.update_or_create(
					tag_name = cell,
				)
		return redirect('admin:index')


def download_tags(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()

	data = Tag.objects.all()

	row_num = 0
	
	column = worksheet.set_column('A:A', 25)
	
	#column header names
	columns = ['Tag Name',]

	#write column headers in sheet
	for col_num in range(len(columns)):
		worksheet.write(row_num, col_num, columns[col_num])
	
	#Writting content on excel sheet
	for my_row in data:	
		row_num = row_num + 1		
		worksheet.write(row_num, 0, my_row.tag_name)
		
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=tag_list.xlsx"

	output.close()

	return response
	

'''
	End Upload Tag Information
'''


'''
	Start upload Sub Theme Information
'''
def sub_theme_upload(request):
	upload_template = "study/sub_theme_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/sub_theme_confirm.html", {"excel_data":excel_data})


def sub_theme_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			got_sub_theme_name = row[0]	
			theme_instance = (row[1])	
			theme_object = Theme.objects.get(theme_name = theme_instance)
			_, object_instance = SubTheme.objects.update_or_create(
				sub_theme_name = got_sub_theme_name,
				theme = theme_object,
			)	
		return redirect('admin:index')


def download_sub_theme_related_template(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()


	data = Theme.objects.all()

	row_num = 0

	
	column = worksheet.set_column('A:A', 45)

	#Writting content on excel sheet
	for my_row in data:		
		
		worksheet.write(row_num, 0, my_row.theme_name)
		row_num = row_num + 1
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=sub_theme_related_template.xlsx"

	output.close()

	return response

'''
	End Upload of Sub Theme Information
'''

'''
	Start Upload of Sub Category Information
'''

def sub_category_upload(request):
	upload_template = "study/sub_category_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Sheet1"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>1:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/sub_category_confirm.html", {"excel_data":excel_data})


def sub_category_upload_confirm(request):
	
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			got_sub_category_name = row[0]	
			category_instance = (row[1])	
			category_object = Category.objects.get(category_name = category_instance)
			_, object_instance = SubCategory.objects.update_or_create(
				sub_category_name = got_sub_category_name,
				category = category_object,
			)	
		return redirect('admin:index')


def download_sub_category_related_template(request):
	output = io.BytesIO()

	workbook = Workbook(output, {'in_memory': True})
	worksheet = workbook.add_worksheet()


	data = Category.objects.all()

	row_num = 0

	
	column = worksheet.set_column('A:A', 45)

	#Writting content on excel sheet
	for my_row in data:		
		
		worksheet.write(row_num, 0, my_row.category_name)
		row_num = row_num + 1
	workbook.close()

	output.seek(0)

	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	response['Content-Disposition'] = "attachment; filename=sub_category_related_template.xlsx"

	output.close()
	return response

'''
	End Upload of Sub Category Information
'''

'''
	Start Study Views
'''

def study_upload(request):
	upload_template = "study/study_upload.html"
	prompt = {}

	if request.method == "GET":
		return render(request, upload_template, prompt)
	else:
		#Capture uploaded file
		excel_file = request.FILES['file_upload']

		#validation to ensure file is excel file
		if not excel_file.name.endswith('.xlsx'):
			messages.error(request, 'This is not an excel file')
		else:
			wb = openpyxl.load_workbook(excel_file)

			# getting a particular sheet by name out of many sheets
			worksheet = wb["Study"]
			#print(worksheet)

			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			x = 1
			for row in worksheet.iter_rows():
				row_data = list()				
				if x>2:
					for cell in row:
						if cell !="":
							row_data.append(str(cell.value))
					excel_data.append(row_data)
				x = x + 1

			request.session['list'] = excel_data
			return render(request, "study/study_confirm.html", {"excel_data":excel_data})

	return render(request, template, context)


def study_upload_confirm(request):
	if request.method == "POST":
		#excel_data = []
		excel_data = request.session['list']

		for row in excel_data:
			got_study_title = row[0]
			got_study_author = row[1]
			got_study_year = row[2]
			country_instance = (row[3])
			region_instance = (row[4])
			resource_instance = (row[5])
			got_quality_data = (row[6])
			got_study_link = (row[7])


			theme_one_instance = (row[8])
			theme_two_instance = (row[9])
			theme_three_instance = (row[10])
			theme_four_instance = (row[11])
			theme_five_instance = (row[12])
			theme_six_instance = (row[13])
			theme_seven_instance = (row[14])
			theme_eight_instance = (row[15])

			category_one_instance = (row[16])
			category_two_instance = (row[17])
			category_three_instance = (row[18])
			category_four_instance = (row[19])
			category_five_instance = (row[20])
			
			try:
				region_object = Region.objects.get(region_name = region_instance)
			except Region.DoesNotExist:
				region_object = None
			try:
				country_object = Country.objects.get(country_name = country_instance)
			except Country.DoesNotExist:
				country_object = None
			try:
				resource_object = Resource.objects.get(resource_name = resource_instance)
			except Resource.DoesNotExist:
				resource_object = None
			
			try:
				theme_one_object = Theme.objects.get(theme_name = theme_one_instance)
			except Theme.DoesNotExist:
				theme_one_object = None
			try:
				theme_two_object = Theme.objects.get(theme_name = theme_two_instance)
			except Theme.DoesNotExist:
				theme_two_object = None
			try:
				theme_three_object = Theme.objects.get(theme_name = theme_three_instance)
			except Theme.DoesNotExist:
				theme_three_object = None
			try:
				theme_four_object = Theme.objects.get(theme_name = theme_four_instance)
			except Theme.DoesNotExist:
				theme_four_object = None
			try:
				theme_five_object = Theme.objects.get(theme_name = theme_five_instance)
			except Theme.DoesNotExist:
				theme_five_object = None
			try:
				theme_six_object = Theme.objects.get(theme_name = theme_six_instance)
			except Theme.DoesNotExist:
				theme_six_object = None
			try:
				theme_seven_object = Theme.objects.get(theme_name = theme_seven_instance)
			except Theme.DoesNotExist:
				theme_seven_object = None
			try:
				theme_eight_object = Theme.objects.get(theme_name = theme_eight_instance)
			except Theme.DoesNotExist:
				theme_eight_object = None
			
			try:
				category_one_object = Category.objects.get(category_name = category_one_instance)
			except Category.DoesNotExist:
				category_one_object = None
			try:
				category_two_object = Category.objects.get(category_name = category_two_instance)
			except Category.DoesNotExist:
				category_two_object = None
			try:
				category_three_object = Category.objects.get(category_name = category_three_instance)
			except Category.DoesNotExist:
				category_three_object = None
			try:
				category_four_object = Category.objects.get(category_name = category_four_instance)
			except Category.DoesNotExist:
				category_four_object = None
			try:
				category_five_object = Category.objects.get(category_name = category_five_instance)
			except Category.DoesNotExist:
				category_five_object = None
			
			object_instance = Study.objects.create(				
				link = got_study_link,
				year = got_study_year,				
				title = got_study_title,				
				quality = got_quality_data,
				author = got_study_author,
				resource = resource_object,
			)	
			if region_object:
				object_instance.region.add(region_object)

			if country_object:
				object_instance.country.add(country_object)

			if theme_one_object is not None:
				object_instance.theme.add(theme_one_object)

			if theme_two_object is not None:
				if theme_one_object is not theme_two_object:
					object_instance.theme.add(theme_two_object)
			
			if theme_three_object is not None:
				if theme_three_object is not theme_one_object:
					if theme_three_object is not theme_two_object:
						object_instance.theme.add(theme_three_object)
			
			if theme_four_object is not None:
				if theme_four_object is not theme_one_object:
					if theme_four_object is not theme_two_object:
						if theme_four_object is not theme_three_object:
							object_instance.theme.add(theme_four_object)

			if theme_five_object is not None:
				if theme_five_object is not theme_one_object:
					if theme_five_object is not theme_two_object:
						if theme_five_object is not theme_three_object:
							if theme_five_object is not theme_four_object:
								object_instance.theme.add(theme_five_object)

			if theme_six_object is not None:
				if theme_six_object is not theme_one_object:
					if theme_six_object is not theme_two_object:
						if theme_six_object is not theme_three_object:
							if theme_six_object is not theme_four_object:
								if theme_six_object is not theme_five_object:
									object_instance.theme.add(theme_six_object)

			if theme_seven_object is not None:
				if theme_seven_object is not theme_one_object:
					if theme_seven_object is not theme_two_object:
						if theme_seven_object is not theme_three_object:
							if theme_seven_object is not theme_four_object:
								if theme_seven_object is not theme_five_object:
									if theme_seven_object is not theme_six_object:
										object_instance.theme.add(theme_seven_object)

			if theme_eight_object is not None:
				if theme_eight_object is not theme_one_object:
					if theme_eight_object is not theme_two_object:
						if theme_eight_object is not theme_three_object:
							if theme_eight_object is not theme_four_object:
								if theme_eight_object is not theme_five_object:
									if theme_eight_object is not theme_six_object:
										if theme_eight_object is not theme_seven_object:
											object_instance.theme.add(theme_eight_object)


			if category_one_object is not None:
				object_instance.category.add(category_one_object)

			if category_two_object is not None:
				if category_one_object is not category_two_object:
					object_instance.category.add(category_two_object)
			
			if category_three_object is not None:
				if category_three_object is not category_one_object:
					if category_three_object is not category_two_object:
						object_instance.category.add(category_three_object)
			
			if category_four_object is not None:
				if category_four_object is not category_one_object:
					if category_four_object is not category_two_object:
						if category_four_object is not category_three_object:
							object_instance.category.add(category_four_object)

			if category_five_object is not None:
				if category_five_object is not category_one_object:
					if category_five_object is not category_two_object:
						if theme_five_object is not category_three_object:
							if theme_five_object is not category_four_object:
								object_instance.category.add(category_five_object)


		return redirect('admin:index')
'''
	End Study Views
'''


'''
Method to download template files
'''
def download_template(request, file_name):	
	file_path = os.path.join(settings.MEDIA_ROOT, file_name)
	if os.path.exists(file_path):
		with open(file_path, 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
			response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
			return response

	
def download_study_related_template(request):
	#output = io.BytesIO()

	#workbook = Workbook(output, {'in_memory': True})
	#worksheet = workbook.add_worksheet()

	output = io.BytesIO()

	wrkbook = Workbook(output, {'in_memory': True})
	

	#Region Worksheet	
	region_wrksheet = wrkbook.add_worksheet("region")

	region_data = Region.objects.all()
	region_row_num = 0
	#Writting content on excel sheet
	for my_row in region_data:
		#region_wrksheet.write(region_row_num, 0, my_row.id)
		region_wrksheet.write(region_row_num, 0, my_row.region_name)
		region_row_num = region_row_num + 1

	country_wrksheet = wrkbook.add_worksheet("country")
	country_data = Country.objects.all()
	country_row_num = 0
	#Writting content on excel sheet
	for my_row in country_data:
		#country_wrksheet.write(country_row_num, 0, my_row.id)
		country_wrksheet.write(country_row_num, 0, my_row.country_name)
		country_row_num = country_row_num + 1
	'''
	quality_wrksheet = wrkbook.add_worksheet("quality")
	quality_data = Quality.objects.all()
	quality_row_num = 0
	#Writting content on excel sheet
	for my_row in quality_data:
		#quality_wrksheet.write(quality_row_num, 0, my_row.id)
		quality_wrksheet.write(quality_row_num, 0, my_row.quality_name)
		quality_row_num = quality_row_num + 1
	'''
	resource_wrksheet = wrkbook.add_worksheet("resource")
	resource_data = Resource.objects.all()
	resource_row_num = 0
	#Writting content on excel sheet
	for my_row in resource_data:
		#resource_wrksheet.write(resource_row_num, 0, my_row.id)
		resource_wrksheet.write(resource_row_num, 0, my_row.resource_name)
		resource_row_num = resource_row_num + 1
	'''
	theme_wrksheet = wrkbook.add_worksheet("theme")
	theme_data = Theme.objects.all()
	theme_row_num = 0
	#Writting content on excel sheet
	for my_row in theme_data:
		#sub_theme_wrksheet.write(sub_theme_row_num, 0, my_row.id)
		theme_wrksheet.write(theme_row_num, 0, my_row.theme_name)
		theme_row_num = theme_row_num + 1

	categroy_wrksheet = wrkbook.add_worksheet("category")
	categroy_data = Category.objects.all()
	categroy_row_num = 0
	#Writting content on excel sheet
	for my_row in categroy_data:
		#sub_categroy_wrksheet.write(sub_categroy_row_num, 0, my_row.id)
		categroy_wrksheet.write(categroy_row_num, 0, my_row.category_name)
		categroy_row_num = categroy_row_num + 1
	'''
	wrkbook.close()

	output.seek(0)

	# content-type of response
	response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

	#decide file name
	response['Content-Disposition'] = "attachment; filename=study_related_templates.xlsx"

	output.close()
		
	return response


'''
	End downlodaing Templates
'''
	